import asyncio
import aiohttp
import pandas as pd
import openpyxl
from openpyxl.styles import PatternFill, Font, Border, Side



def split_comment(comment, max_length):
    """Split a comment into parts where each part is within the max_length."""
    words = comment.split()
    parts = []
    current_part = ""

    for word in words:
        if len(current_part) + len(word) + 1 <= max_length:
            current_part += (word + " ")
        else:
            parts.append(current_part.strip())
            current_part = word + " "

    if current_part:
        parts.append(current_part.strip())

    return parts

def format_comments(comments, max_length=5120):
    formatted_comments = []
    comment_id = 1

    for comment in comments:
        if len(comment) <= max_length:
            formatted_comments.append({
                "id": str(comment_id),
                "language": "en",
                "text": comment
            })
            comment_id += 1
        else:
            split_parts = split_comment(comment, max_length)
            for i, part in enumerate(split_parts, start=1):
                formatted_comments.append({
                    "id": f"{comment_id}.{i}",
                    "language": "en",
                    "text": part
                })
            comment_id += 1

    return formatted_comments

# Example usage
comments = [
    "A very long comment that might exceed the limit..." * 250,
    "Another comment here.",
    "Yet another comment."
]

formatted_comments = format_comments(comments)




url = "/language/:analyze-text?api-version=2023-04-01"
# Define the request headers
headers = {
        "Content-Type": "application/json",
        "Ocp-Apim-Subscription-Key": 
}


async def fetch_sentiment(session, url, headers, batch):
    payload = {
        "kind": "SentimentAnalysis",
        "parameters": {
            "modelVersion": "latest",
            "opinionMining": "True"
        },
        "analysisInput": {
            "documents": batch
        }
    }
    async with session.post(url, headers=headers, json=payload) as response:
        return await response.json()

async def main(comments, batch_size=10):
    formatted_comments = format_comments(comments)
    
    stats = {"api_calls": 0, "documents_processed": 0}

    async with aiohttp.ClientSession() as session:
        tasks = []
        for i in range(0, len(formatted_comments), batch_size):
            batch = formatted_comments[i:i + batch_size]
            task = fetch_sentiment(session, url, headers, batch)
            tasks.append(task)
            stats["api_calls"] += 1
            stats["documents_processed"] += len(batch)
        responses = await asyncio.gather(*tasks)

    return responses, stats



comments = [
    "A very long comment that might exceed the limit..." * 250,
    "Another comment here.",
    "Yet another comment."
]

responses, stats = await main(comments)

def interpret_confidence_scores(positive, neutral, negative):
    # Define thresholds for strong and weak sentiment
    strong_threshold = 0.75
    weak_threshold = 0.5

    if positive >= strong_threshold:
        return "Strong Positive"
    elif positive >= weak_threshold:
        return "Weak Positive"
    elif negative >= strong_threshold:
        return "Strong Negative"
    elif negative >= weak_threshold:
        return "Weak Negative"
    else:
        return "Neutral"




def aggregate_results(responses):
    aggregated_results = {}
    for response in responses:
        for doc in response['results']['documents']:
            main_id, _, sub_id = doc['id'].partition('.')
            if main_id not in aggregated_results:
                aggregated_results[main_id] = {
                    "sentiment": [],
                    "confidenceScores": {"positive": [], "neutral": [], "negative": []},
                    "positive_opinions": [],
                    "negative_opinions": []
                }
            
            aggregated_results[main_id]['sentiment'].append(doc['sentiment'])
            for key in doc['confidenceScores']:
                aggregated_results[main_id]['confidenceScores'][key].append(doc['confidenceScores'][key])
            
            for sentence in doc['sentences']:
                for target in sentence.get('targets', []):
                    target_text = target['text']
                    assessments = []
                    for relation in target.get('relations', []):
                        ref = relation['ref']
                        path = ref.split('/')
                        assessment_index = int(path[-1])
                        assessment = sentence['assessments'][assessment_index]
                        assessments.append(assessment['text'])
                    opinion = f"{target_text} - {'; '.join(assessments)}"
                    if target['sentiment'] == 'positive':
                        aggregated_results[main_id]['positive_opinions'].append(opinion)
                    elif target['sentiment'] == 'negative':
                        aggregated_results[main_id]['negative_opinions'].append(opinion)

    return aggregated_results

# Assuming responses is the list of responses from the API
aggregated_results = aggregate_results(responses)

# Create a DataFrame
df_data = []
for comment_id, data in aggregated_results.items():
    avg_confidence = {key: sum(values)/len(values) for key, values in data['confidenceScores'].items()}
    sentiment_label = interpret_confidence_scores(avg_confidence['positive'], avg_confidence['neutral'], avg_confidence['negative'])

    row = {
        "id": comment_id,
        "combined_sentiment": ' '.join(data['sentiment']),
        "average_confidence": avg_confidence,
        "positive_opinions": ' | '.join(data['positive_opinions']),
        "negative_opinions": ' | '.join(data['negative_opinions']),
        "sentiment_label": sentiment_label
    }
    df_data.append(row)

df = pd.DataFrame(df_data)
df


def apply_color(worksheet, start_row, start_col, df):
    color_map = {
        "Strong Positive": "00FF00",  # Green
        "Weak Positive": "CCFFCC",   # Light Green
        "Neutral": "FFFF99",         # Yellow
        "Weak Negative": "FFCCCC",   # Light Red
        "Strong Negative": "FF0000"  # Red
    }

    for index, value in enumerate(df['sentiment_label'], start=start_row):
        cell = worksheet.cell(row=index, column=start_col)
        cell.fill = openpyxl.styles.PatternFill(start_color=color_map[value], end_color=color_map[value], fill_type="solid")

# Function to style the header
def style_header(worksheet, df):
    header_fill = PatternFill(start_color="ADD8E6", end_color="ADD8E6", fill_type="solid")
    font = Font(bold=True)
    border = Border(left=Side(style='thin'), 
                    right=Side(style='thin'), 
                    top=Side(style='thin'), 
                    bottom=Side(style='thin'))

    for col in range(1, len(df.columns) + 1):
        cell = worksheet.cell(row=1, column=col)
        cell.fill = header_fill
        cell.font = font
        cell.border = border

def set_column_width(worksheet, column_letter, width):
    worksheet.column_dimensions[column_letter].width = width


# Save DataFrame to Excel
excel_file = "output.xlsx"
with pd.ExcelWriter(excel_file, engine='openpyxl') as writer:
    df.to_excel(writer, index=False)

    # Apply coloring and styling
    workbook = writer.book
    worksheet = writer.sheets['Sheet1']
    apply_color(worksheet, start_row=2, start_col=df.columns.get_loc('sentiment_label') + 1, df=df)
    style_header(worksheet, df)

    # Set width for 'negative mined opinions' column
    negative_opinions_col = openpyxl.utils.get_column_letter(df.columns.get_loc('negative_opinions') + 1)
    set_column_width(worksheet, negative_opinions_col, 20)  # Adjust the width as needed
# Inform the user
print(f"Excel file '{excel_file}' has been created.")









